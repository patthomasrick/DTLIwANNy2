from openpyxl import Workbook
from openpyxl import utils
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

from xml2trainingdata import DEFAULT_TRAINING_FILE, DEFAULT_SAMPLE_SAVE_FILE, load_leaf_samples
from annmanager import MAX_ITERATIONS, DESIRED_ERROR

"""
output-to-xls.py: Provides the tools to save the outputs from the ANN to an excel spreadsheet.
"""

__author__ = "Patrick Thomas"
__credits__ = ["Patrick Thomas", "Rick Fisher"]
__version__ = "1.0.0"
__date__ = "12/15/16"
__maintainer__ = "Patrick Thomas"
__email__ = "pthomas@mail.swvgs.us"
__status__ = "Development"

# global vars

# filename is in the following format:
# output-YYYYMMDD-HHMM

# 2006  # year
# 11  # month
# 21  # day
# 16  # hour
# 30  # minute
# 0  # second
# 1  # weekday (0 = Monday)
# 325  # number of days since 1st January
# -1  # dst - method tzinfo.dst() returned None

# biuld the timestamp for the filename
dt = datetime.now()
time_tuple = dt.timetuple()
y, mon, d, h, m, s, wd, td, dst = time_tuple

TIMESTAMP = '{0}{1}{2}-{3}{4}'.format(
    str(y).zfill(4),
    str(mon).zfill(2),
    str(d).zfill(2),
    str(h).zfill(2),
    str(m).zfill(2),
)

DEFAULT_WB_FILENAME = 'output/data-out-{0}.xlsx'.format(TIMESTAMP)
print('Saving workbook to {0}.'.format(DEFAULT_WB_FILENAME))


def output(filename=DEFAULT_WB_FILENAME):
    """
    12/15/16

    :param filename:
    :return:
    """
    save_ann_output_to_excel(filename, leaf_data=None, ann_output=None)


def save_leaf_data_to_excel(filename, leaf_data):
    pass


def save_ann_output_to_excel(filename, run_output):
    # create a workbook
    wb = Workbook()

    # get the current spreadsheet
    ws = wb.active
    ws.title = 'DTLIANN Output {0}'.format(TIMESTAMP)

    # set simple information cells, such as title and neuron numbers, etc.
    # title
    ws['A1'] = 'DTLIANN Output'
    # time
    ws['B1'] = TIMESTAMP

    # read the first line of the training file to get neuron numbers
    with open(DEFAULT_TRAINING_FILE, 'r') as train_file:
        header = train_file.readline()
        train_file.close()
    total, num_input, num_output = header.split(' ')
    num_input = int(num_input)
    num_output = int(num_output)

    num_hidden = (2 * (num_output + num_input) / 3)

    # neuron numbers
    ws['A2'] = 'Input neurons'
    ws['B2'] = num_input
    ws['A3'] = 'Output neurons'
    ws['B3'] = num_output
    ws['A4'] = 'Hidden neurons'
    ws['B4'] = num_hidden

    ws['D2'] = 'Max. Iterations'
    ws['E2'] = MAX_ITERATIONS
    ws['D3'] = 'Desired MSE'
    ws['E3'] = DESIRED_ERROR

    # start the actual data portion
    ws['A8'] = 'Species Ran'

    # ws['B7'] = 'ANN Output'

    ws['C6'] = 'ANN\'s Guess'

    # LOAD DATA TO BE SAVED
    # leaf data for leaf names
    all_leaves, train_leaves, run_leaves = load_leaf_samples(DEFAULT_SAMPLE_SAVE_FILE)
    all_species = [k for k in train_leaves.keys()]
    all_species = sorted(all_species)

    species_pop = all_species.copy()

    # save to cell range
    header_range = 'C7:{0}7'.format(utils.get_column_letter(2 + len(all_species)))
    for row in ws[header_range]:
        for cell in row:
            # save leaf species to cell as well as abbreviate
            bin_nom = all_species.pop(0)
            g, s = bin_nom.split(' ')
            g_abbrev = g[0] + '.'
            cell.value = '{0} {1}'.format(g_abbrev, s)

    # save the actual ANN output
    # sort run data dict by species
    sorted_output = sorted(run_output, key=lambda k: k['species'])

    # get range
    species_range = ws['B8:B{0}'.format(7 + len(run_output))]
    data_range = ws['C8:{0}{1}'.format(utils.get_column_letter(2 + len(train_leaves)), 7 + len(run_output))]

    # write all species first
    for i, row in enumerate(species_range):
        for cell in row:
            # save leaf species to cell as well as abbreviate
            g, s = sorted_output[i]['species'].split(' ')
            g_abbrev = g[0] + '.'
            cell.value = '{0} {1}'.format(g_abbrev, s)

    # now write all ANN outputs
    for i, row in enumerate(data_range):
        # print(row)
        for j, cell in enumerate(row):
            # print(sorted_output[i]['ann'][j])
            cell.value = ((sorted_output[i]['ann'][j]/2)+0.5)
            cell.style = 'Percent'

    # for col_count, col in data_range:
    #     for row_count, cell in enumerate(col):
    #         print(str(run_output[col_count]['ann'][row_count - 1]))
    #         if row_count is 0:
    #             cell.value = run_output[col_count]['species']
    #         else:
    #             cell.value = str(run_output[col_count]['ann'][row_count - 1])

    # SAVE THE FILE TO THE DISK

    # save the workbook
    wb.save(filename)


def save_ann_output_to_excel_one_file(filename, worksheet_name, run_output, last_row=8, ws_sum_line=8):
    # create a workbook
    # wb = None
    # ws = None
    # ws_main = None

    new_wb = False

    try:
        # create new sheet if workbook exists
        wb = load_workbook(filename)

        # get the summary sheet
        ws_sum = wb.get_sheet_by_name('Summary')
        
        ws_sum_avg = wb.get_sheet_by_name('Summary 2')

        # create the individual run sheet
        ws = wb.create_sheet(worksheet_name)

    except FileNotFoundError:  # if the workbook doesn't exist, create it
        # create new workbook if not
        wb = Workbook()

        # create the summary sheets
        ws_sum = wb.active
        ws_sum.title = 'Summary'
        
        ws_sum_avg = wb.create_sheet('Summary 2')

        # create the individual run sheet
        ws = wb.create_sheet(worksheet_name)

        new_wb = True

    ##################################
    #
    #       INDIVIDUAL SHEETS
    #
    ##################################

    # set simple information cells, such as title and neuron numbers, etc.
    # title
    ws['A1'] = 'DTLIANN Output'
    # time
    ws['B1'] = TIMESTAMP

    # read the first line of the training file to get neuron numbers
    with open(DEFAULT_TRAINING_FILE, 'r') as train_file:
        header = train_file.readline()
        train_file.close()
    total, num_input, num_output = header.split(' ')
    num_input = int(num_input)
    num_output = int(num_output)

    # caculate the number of hidden neurons
    num_hidden = int(2 * (num_output + num_input) / 3)

    # save the number of neurons to the file
    ws['A2'] = 'Input neurons'
    ws['B2'] = num_input
    ws['A3'] = 'Output neurons'
    ws['B3'] = num_output
    ws['A4'] = 'Hidden neurons'
    ws['B4'] = num_hidden

    ws['D2'] = 'Max. Iterations'
    ws['E2'] = MAX_ITERATIONS
    ws['D3'] = 'Desired MSE'
    ws['E3'] = DESIRED_ERROR

    # start the actual data portion
    ws['A8'] = 'Species Ran'

    # ws['B7'] = 'ANN Output'

    ws['C6'] = 'ANN\'s Guess'

    # LOAD DATA TO BE SAVED
    # leaf data for leaf names
    all_leaves, train_leaves, run_leaves = load_leaf_samples(DEFAULT_SAMPLE_SAVE_FILE)
    all_species = [k for k in train_leaves.keys()]
    all_species = sorted(all_species)

    species_pop = all_species.copy()

    # save to cell range
    header_range = 'C7:{0}7'.format(utils.get_column_letter(2 + len(all_species)))
    for row in ws[header_range]:
        for cell in row:
            # save leaf species to cell as well as abbreviate
            bin_nom = all_species.pop(0)
            g, s = bin_nom.split(' ')
            g_abbrev = g[0] + '.'
            cell.value = '{0} {1}'.format(g_abbrev, s)

    # save the actual ANN output
    # sort run data dict by species
    sorted_output = sorted(run_output, key=lambda k: k['species'])

    # get range to save leaf data to
    species_range = ws['B8:B{0}'.format(7 + len(run_output))]
    data_range = ws['C8:{0}{1}'.format(utils.get_column_letter(2 + len(train_leaves)), 7 + len(run_output))]

    # write all species first
    for i, row in enumerate(species_range):
        for cell in row:

            # save leaf species names to cells as well as abbreviate genus
            g, s = sorted_output[i]['species'].split(' ')
            g_abbrev = g[0] + '.'
            cell.value = '{0} {1}'.format(g_abbrev, s)

    # now write all ANN outputs
    for i, row in enumerate(data_range):
        for j, cell in enumerate(row):

            # write cell's value and format it to a percent
            cell.value = ((sorted_output[i]['ann'][j]/2)+0.5)
            cell.style = 'Percent'

    # save the workbook
    wb.save(filename)

    ##################################
    #
    #         SUMMARY SHEET
    #
    ##################################

    # do this if new workbook
    if new_wb:
        # SUMMARY
        # set simple information cells, such as title and neuron numbers, etc.
        # title
        ws_sum['A1'] = 'DTLIANN Data Summary'
        # time
        ws_sum['B1'] = TIMESTAMP

        ws_sum['D1'] = 'Max. Iterations'
        ws_sum['E1'] = MAX_ITERATIONS
        ws_sum['D2'] = 'Desired MSE'
        ws_sum['E2'] = DESIRED_ERROR

        # data portion
        ws_sum['A8'] = 'Species Ran'

        ws_sum['C7'] = 'Num. of Species'
        ws_sum['D7'] = 'Certainty'
        ws_sum['E7'] = 'Correct?'
        
        # SUMMARY SUMMARY
        # set simple information cells, such as title and neuron numbers, etc.
        # title
        ws_sum_avg['A1'] = 'DTLIANN Data Summary 2'
        # time
        ws_sum_avg['B1'] = TIMESTAMP

        ws_sum_avg['D1'] = 'Max. Iterations'
        ws_sum_avg['E1'] = MAX_ITERATIONS
        ws_sum_avg['D2'] = 'Desired MSE'
        ws_sum_avg['E2'] = DESIRED_ERROR

        # data portion
        ws_sum_avg['A8'] = 'Trials'
        ws_sum_avg['B6'] = 'Num'
        ws_sum_avg['B7'] = 'Species'
        ws_sum_avg['C6'] = 'Avg'
        ws_sum_avg['C7'] = 'Correctness'

    # get range to save leaf data to
    range_numbers = (last_row, last_row + len(sorted_output) - 1)
    species_range = ws_sum['B{0}:B{1}'.format(
        range_numbers[0],
        range_numbers[1])]
    data_range = ws_sum['C{0}:E{1}'.format(
        range_numbers[0],
        range_numbers[1])]
    avg_range = ['E{0}'.format(range_numbers[0]),
                 'E{0}'.format(range_numbers[1])]

    # write species names
    for i, row in enumerate(species_range):
        for cell in row:
            # save leaf species names to cells as well as abbreviate genus
            g, s = sorted_output[i]['species'].split(' ')
            cell.value = '{0}. {1}'.format(g[0], s)

    # write average
    avg_cell_address = 'F{0}'.format(range_numbers[0])
    ws_sum[avg_cell_address] = '=AVERAGE({0}:{1})'.format(avg_range[0], avg_range[1])
    ws_sum[avg_cell_address].style = 'Percent'

    ######################################
    #
    # write lines to summary sheet 2
    #
    #######################################
    # get the column and cell addresses first
    col_0 = utils.get_column_letter(3*num_output - 5)
    col_1 = utils.get_column_letter(3*num_output - 4)
    col_2 = utils.get_column_letter(3*num_output - 3)
    cell_1 = '{0}{1}'.format(col_1, ws_sum_line)
    cell_2 = '{0}{1}'.format(col_2, ws_sum_line)

    # always write headers
    # for reference, see the first-time-only code above
    ws_sum_avg[col_0 + '8'] = 'Trials'
    ws_sum_avg[col_1 + '6'] = 'AVG'
    ws_sum_avg[col_1 + '7'] = 'Num. Species'
    ws_sum_avg[col_2 + '7'] = 'Correctness'

    # then write to cells
    ws_sum_avg[cell_1] = num_output
    ws_sum_avg[cell_2] = '=Summary!{0}'.format(avg_cell_address)
    ws_sum_avg[cell_2].style = 'Percent'

    # now write all ANN outputs to run specific sheet
    # get a list of all species to aid with finding the ANN's choice
    output_species_list = []
    for d in sorted_output:
        output_species_list.append(d['species'])

    output_species_list = sorted(list(set(output_species_list)))

    for i, row in enumerate(data_range):
        for j, cell in enumerate(row):
            actual_species = sorted_output[i]['species']
            ann_output = sorted_output[i]['ann']

            ann_out_dict = {species: ann_output[i] for i, species in enumerate(output_species_list)}

            # if at the first column, num of species
            if j == 0:
                cell.value = num_output

            # else if at the second column, certainty
            elif j == 1:
                cell.value = ann_out_dict[actual_species]/2 + 0.5
                cell.style = 'Percent'

            # else if at the third column, correctness
            elif j == 2:
                # get the ANN's choice for the entry
                max_val = max(ann_output)
                # get all values that are maximum
                indices = [i for i, val in enumerate(ann_output) if val == max_val]

                # if there are multiple maximums, fail
                if len(indices) > 1:
                    cell.value = 0

                # else there is only 1 maximum
                else:
                    # if the species matches the actual species:
                    if ann_out_dict[actual_species] == max_val:
                        cell.value = 1
                    else:  # else, fail
                        cell.value = 0

    # save workbook
    wb.save(filename)

    # return the last cell used to resume inserting values
    return range_numbers[1] + 1, ws_sum_line + 1




if __name__ == '__main__':
    output()
